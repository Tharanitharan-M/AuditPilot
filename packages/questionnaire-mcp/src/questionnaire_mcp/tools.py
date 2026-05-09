"""Questionnaire tools (Sprint 7 chunks 7.2-7.5).

Pure-function tools used by the MCP server and (via direct import) by the
``apps/api`` worker. No network calls, no environment access — the API layer
is responsible for fetching the XLSX from R2 and writing the assembled file
back to R2.

Refs: PLAN.md chunks 7.2-7.5; ADR-0005; system-design 3.4, 11.7.
"""

from __future__ import annotations

import re
from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.workbook import Workbook

from questionnaire_mcp.schemas import (
    Answer,
    AnswerType,
    AssembleResult,
    Citation,
    Cluster,
    ClusterResult,
    MetadataResult,
    ParsedQuestionnaire,
    Question,
    QuestionDomain,
    QuestionMetadata,
    QuestionnaireFormat,
)

# ---------------------------------------------------------------------------
# Domain inference: keyword-driven classifier for SIG-Lite questions.
# ---------------------------------------------------------------------------

_DOMAIN_KEYWORDS: dict[QuestionDomain, tuple[str, ...]] = {
    "access_control": (
        "access",
        "authentication",
        "authorization",
        "mfa",
        "multi-factor",
        "least privilege",
        "rbac",
        "sso",
        "password",
        "privilege",
    ),
    "asset_management": (
        "asset",
        "inventory",
        "configuration management",
        "cmdb",
        "lifecycle",
    ),
    "business_resiliency": (
        "business continuity",
        "disaster recovery",
        "rto",
        "rpo",
        "resilien",
        "backup",
    ),
    "compliance": (
        "compliance",
        "regulator",
        "internal audit",
        "readiness assessment",
        "policy",
        "standard",
        "framework",
    ),
    "data_handling": (
        "data classification",
        "data handling",
        "data retention",
        "data loss",
        "dlp",
        "encryption",
        "tokeni",
        "pii",
        "personal data",
        "phi",
    ),
    "endpoint_security": (
        "endpoint",
        "edr",
        "antivirus",
        "device",
        "laptop",
        "workstation",
        "mdm",
    ),
    "human_resources": (
        "background check",
        "hiring",
        "termination",
        "offboarding",
        "onboarding",
        "employee",
        "personnel",
    ),
    "incident_response": (
        "incident",
        "breach",
        "irp",
        "playbook",
        "post-mortem",
        "forensic",
    ),
    "network_security": (
        "firewall",
        "network",
        "vpn",
        "ids",
        "ips",
        "dns",
        "tls",
        "vlan",
        "segment",
    ),
    "operations_management": (
        "monitoring",
        "logging",
        "siem",
        "patch",
        "vulnerability",
        "ticket",
        "change management",
    ),
    "risk_management": (
        "risk",
        "threat",
        "assessment",
        "control",
        "governance",
        "framework",
    ),
    "third_party_management": (
        "vendor",
        "third party",
        "third-party",
        "subprocessor",
        "supplier",
        "contract",
    ),
    "training_and_awareness": (
        "training",
        "awareness",
        "phishing",
        "education",
        "security training",
    ),
}


def _classify_domain(text: str) -> QuestionDomain:
    """Pick the best-matching SIG-Lite domain for a question by keyword overlap."""
    if not text:
        return "uncategorized"
    lowered = text.lower()
    best: tuple[QuestionDomain, int] = ("uncategorized", 0)
    for domain, keywords in _DOMAIN_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in lowered)
        if score > best[1]:
            best = (domain, score)
    return best[0]


# ---------------------------------------------------------------------------
# Answer-type inference: detect yes/no, free text, etc. from the question stem.
# ---------------------------------------------------------------------------

_YES_NO_PARTIAL_RE = re.compile(r"\b(partial|partially)\b", re.IGNORECASE)
_YES_NO_RE = re.compile(r"^\s*(do|does|is|are|have|has|can|will|did)\b", re.IGNORECASE)
_NUMERIC_RE = re.compile(r"\b(how many|number of|count of)\b", re.IGNORECASE)
_DATE_RE = re.compile(r"\b(when did|date of|last reviewed|since when)\b", re.IGNORECASE)


def _infer_answer_type(text: str) -> AnswerType:
    """Heuristic answer type from the question stem."""
    if not text:
        return "unknown"
    if _NUMERIC_RE.search(text):
        return "numeric"
    if _DATE_RE.search(text):
        return "date"
    if _YES_NO_PARTIAL_RE.search(text):
        return "yes_no_partial"
    if _YES_NO_RE.search(text):
        return "yes_no"
    if len(text) > 120:
        return "free_text"
    return "free_text"


# ---------------------------------------------------------------------------
# parse_xlsx
# ---------------------------------------------------------------------------


def _detect_format(workbook: Workbook) -> QuestionnaireFormat:
    """Heuristic SIG-Lite detection: a 'SIG' or 'SIG-Lite' sheet name."""
    for sheet_name in workbook.sheetnames:
        lowered = sheet_name.lower()
        if "sig" in lowered:
            return "sig-lite"
    return "custom"


def _is_question_text(value: object) -> bool:
    """Decide whether a cell value looks like a question."""
    if not isinstance(value, str):
        return False
    stripped = value.strip()
    if len(stripped) < 8:
        return False
    if stripped.endswith("?"):
        return True
    # SIG-Lite often phrases prompts as imperative statements.
    if _YES_NO_RE.match(stripped):
        return True
    return len(stripped) >= 24 and not stripped.endswith(":")


def _question_id_from_ref(sheet: str, row: int, column: int) -> str:
    """Stable id: sheet!RxCy."""
    return f"{sheet}!R{row}C{column}"


def parse_xlsx(file_uri: str) -> ParsedQuestionnaire:
    """Parse a SIG-Lite or custom XLSX into a list of questions.

    The SIG-Lite v2026 fixture used in the eval suite is expected to parse
    to exactly 128 questions (FR-030).
    """
    path = Path(file_uri)
    if not path.exists():
        raise FileNotFoundError(f"questionnaire file not found: {file_uri}")

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    fmt = _detect_format(workbook)
    questions: list[Question] = []
    current_section = ""

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row:
                continue
            # Heuristic: section heading is a single-cell row in column 1.
            non_empty = [c for c in row if c not in (None, "")]
            if len(non_empty) == 1 and isinstance(non_empty[0], str):
                stripped = non_empty[0].strip()
                if 0 < len(stripped) <= 80 and not stripped.endswith("?"):
                    current_section = stripped
                    continue

            for col_idx, cell in enumerate(row, start=1):
                if not _is_question_text(cell):
                    continue
                question_text = str(cell).strip()
                # Answer column: prefer column 3 if cell is in column 2 (typical SIG-Lite),
                # otherwise the cell immediately to the right.
                answer_column = col_idx + 1 if col_idx >= 2 else 3
                qid = _question_id_from_ref(sheet_name, row_idx, answer_column)
                questions.append(
                    Question(
                        id=qid,
                        sheet=sheet_name,
                        row=row_idx,
                        column=answer_column,
                        section=current_section,
                        domain=_classify_domain(question_text),
                        text=question_text,
                        answer_type=_infer_answer_type(question_text),
                    )
                )
    workbook.close()

    return ParsedQuestionnaire(
        file_uri=file_uri,
        format=fmt,
        sheet_count=len(workbook.sheetnames),
        question_count=len(questions),
        questions=questions,
    )


# ---------------------------------------------------------------------------
# cluster_questions
# ---------------------------------------------------------------------------

_DOMAIN_LABELS: dict[QuestionDomain, str] = {
    "access_control": "Access Control",
    "asset_management": "Asset Management",
    "business_resiliency": "Business Resiliency",
    "compliance": "Compliance",
    "data_handling": "Data Handling",
    "endpoint_security": "Endpoint Security",
    "human_resources": "Human Resources",
    "incident_response": "Incident Response",
    "network_security": "Network Security",
    "operations_management": "Operations Management",
    "risk_management": "Risk Management",
    "third_party_management": "Third-Party Management",
    "training_and_awareness": "Training and Awareness",
    "uncategorized": "Uncategorized",
}


def cluster_questions(parsed: ParsedQuestionnaire | dict) -> ClusterResult:
    """Group parsed questions by SIG-Lite domain.

    The SIG-Lite v2026 fixture is expected to cluster into 12 +/- 2 groups
    (FR-031).
    """
    if isinstance(parsed, dict):
        parsed = ParsedQuestionnaire.model_validate(parsed)

    buckets: dict[QuestionDomain, list[str]] = {}
    for q in parsed.questions:
        buckets.setdefault(q.domain, []).append(q.id)

    clusters: list[Cluster] = []
    for domain, qids in sorted(buckets.items(), key=lambda kv: kv[0]):
        clusters.append(
            Cluster(
                id=f"cluster_{domain}",
                domain=domain,
                label=_DOMAIN_LABELS.get(domain, domain.replace("_", " ").title()),
                question_ids=qids,
                size=len(qids),
            )
        )
    return ClusterResult(cluster_count=len(clusters), clusters=clusters)


# ---------------------------------------------------------------------------
# extract_question_metadata
# ---------------------------------------------------------------------------


def extract_question_metadata(parsed: ParsedQuestionnaire | dict) -> MetadataResult:
    """Project per-question metadata (section, domain, answer type)."""
    if isinstance(parsed, dict):
        parsed = ParsedQuestionnaire.model_validate(parsed)

    items: list[QuestionMetadata] = [
        QuestionMetadata(
            question_id=q.id,
            section=q.section,
            domain=q.domain,
            answer_type=q.answer_type,
        )
        for q in parsed.questions
    ]
    return MetadataResult(metadata_count=len(items), items=items)


# ---------------------------------------------------------------------------
# assemble_filled_xlsx
# ---------------------------------------------------------------------------

FLAG_THRESHOLD = 0.70
"""Cells with confidence below this threshold are flagged for human review."""


def _coerce_answers(answers: Iterable[Answer | dict]) -> list[Answer]:
    out: list[Answer] = []
    for a in answers:
        if isinstance(a, dict):
            out.append(Answer.model_validate(a))
        else:
            out.append(a)
    return out


def _format_citation_comment(citations: list[Citation]) -> str:
    if not citations:
        return ""
    parts: list[str] = []
    for idx, cit in enumerate(citations, start=1):
        line = f"[{idx}] evidence_id={cit.evidence_id}"
        if cit.snippet:
            line += f"\n    {cit.snippet[:240]}"
        if cit.source_uri:
            line += f"\n    source: {cit.source_uri}"
        parts.append(line)
    return "AuditPilot citations:\n" + "\n".join(parts)


def assemble_filled_xlsx(
    answers: Iterable[Answer | dict],
    source_uri: str,
    output_uri: str,
    *,
    flagged_column_offset: int = 1,
) -> AssembleResult:
    """Write answers back into the source XLSX as inline strings with citation comments.

    - Always writes ``xl_inline_string`` (never a formula).
    - Citation evidence ids are attached as a cell comment.
    - For confidence below ``FLAG_THRESHOLD``, a ``Flagged`` mark is written
      in the cell ``flagged_column_offset`` columns to the right.
    - The original formatting and column widths are preserved by openpyxl
      when loaded with ``read_only=False``.
    """
    src_path = Path(source_uri)
    if not src_path.exists():
        raise FileNotFoundError(f"source questionnaire not found: {source_uri}")

    workbook = load_workbook(filename=str(src_path))
    answer_list = _coerce_answers(answers)
    flagged = 0
    written = 0

    for ans in answer_list:
        if ans.sheet not in workbook.sheetnames:
            continue
        sheet = workbook[ans.sheet]
        # openpyxl uses 1-based row and column.
        cell = sheet.cell(row=ans.row, column=ans.column)
        # Force inline string (never a formula).
        text = ans.text or ""
        if text.startswith("="):
            text = " " + text
        cell.value = text
        cell.data_type = "s"  # inline string
        comment_text = _format_citation_comment(ans.citations)
        if comment_text:
            cell.comment = Comment(comment_text, "AuditPilot")
        if ans.confidence < FLAG_THRESHOLD:
            flag_cell = sheet.cell(row=ans.row, column=ans.column + flagged_column_offset)
            flag_cell.value = "Flagged"
            flag_cell.data_type = "s"
            flagged += 1
        written += 1

    output_path = Path(output_uri)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))
    workbook.close()

    return AssembleResult(
        output_uri=str(output_path),
        answers_written=written,
        flagged_count=flagged,
        sheet_count=len(workbook.sheetnames),
    )


__all__ = [
    "FLAG_THRESHOLD",
    "assemble_filled_xlsx",
    "cluster_questions",
    "extract_question_metadata",
    "parse_xlsx",
]
