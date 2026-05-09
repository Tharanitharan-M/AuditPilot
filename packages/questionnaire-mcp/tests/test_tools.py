"""Tool tests for questionnaire-mcp (Sprint 7 chunks 7.2-7.5)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from questionnaire_mcp.schemas import (
    Answer,
    Citation,
    ParsedQuestionnaire,
)
from questionnaire_mcp.tools import (
    FLAG_THRESHOLD,
    assemble_filled_xlsx,
    cluster_questions,
    extract_question_metadata,
    parse_xlsx,
)


class TestParseXlsx:
    def test_parses_sig_lite_fixture(self, sig_lite_fixture: Path) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        assert parsed.format == "sig-lite"
        assert parsed.sheet_count == 1
        # Six section headers, twenty-four questions in fixture.
        assert parsed.question_count == 24
        assert all(q.text.endswith("?") or "do you" in q.text.lower() for q in parsed.questions)

    def test_classifies_domains(self, sig_lite_fixture: Path) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        domains = {q.domain for q in parsed.questions}
        # Expect at least these domains to be present.
        assert "access_control" in domains
        assert "data_handling" in domains
        assert "incident_response" in domains
        assert "network_security" in domains
        assert "third_party_management" in domains
        assert "training_and_awareness" in domains

    def test_assigns_sections(self, sig_lite_fixture: Path) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        sections = {q.section for q in parsed.questions}
        assert "Access Control" in sections
        assert "Incident Response" in sections

    def test_answer_column_after_question_column(self, sig_lite_fixture: Path) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        # Questions live in column 2; answers in column 3.
        assert all(q.column == 3 for q in parsed.questions)

    def test_missing_file_raises(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            parse_xlsx(str(tmp_path / "nope.xlsx"))


class TestClusterQuestions:
    def test_clusters_by_domain(self, sig_lite_fixture: Path) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        result = cluster_questions(parsed)
        assert result.cluster_count >= 4
        assert result.cluster_count <= 14
        # Sum of cluster sizes equals question count.
        assert sum(c.size for c in result.clusters) == parsed.question_count
        # Each cluster id matches the cluster_<domain> pattern.
        assert all(c.id.startswith("cluster_") for c in result.clusters)

    def test_accepts_dict_input(self, sig_lite_fixture: Path) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        result = cluster_questions(parsed.model_dump())
        assert result.cluster_count > 0


class TestExtractQuestionMetadata:
    def test_one_metadata_per_question(self, sig_lite_fixture: Path) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        result = extract_question_metadata(parsed)
        assert result.metadata_count == parsed.question_count
        assert len(result.items) == parsed.question_count

    def test_metadata_preserves_domain_and_section(self, sig_lite_fixture: Path) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        result = extract_question_metadata(parsed)
        by_id = {m.question_id: m for m in result.items}
        for q in parsed.questions:
            assert by_id[q.id].domain == q.domain
            assert by_id[q.id].section == q.section


class TestAssembleFilledXlsx:
    def test_writes_inline_strings_and_comments(
        self, sig_lite_fixture: Path, tmp_path: Path
    ) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        # Draft an answer for each question — half above threshold, half below.
        answers: list[Answer] = []
        for idx, q in enumerate(parsed.questions):
            confidence = 0.92 if idx % 2 == 0 else 0.45
            answers.append(
                Answer(
                    question_id=q.id,
                    sheet=q.sheet,
                    row=q.row,
                    column=q.column,
                    text=f"Yes — control {idx}",
                    confidence=confidence,
                    citations=[
                        Citation(
                            evidence_id=f"ev_{idx}",
                            snippet="readiness sample evidence",
                            source_uri=f"r2://bucket/ev_{idx}.json",
                        )
                    ],
                )
            )

        out_path = tmp_path / "filled.xlsx"
        result = assemble_filled_xlsx(answers, str(sig_lite_fixture), str(out_path))

        assert result.answers_written == len(answers)
        # Half were below threshold and should be flagged.
        assert result.flagged_count == len([a for a in answers if a.confidence < FLAG_THRESHOLD])

        # Reload and verify the cells.
        wb = load_workbook(str(out_path))
        ws = wb.active
        a0 = answers[0]
        cell = ws.cell(row=a0.row, column=a0.column)
        assert cell.value == a0.text
        # Should not be a formula.
        assert not str(cell.value).startswith("=")
        assert cell.comment is not None
        assert "ev_0" in cell.comment.text

    def test_formula_text_is_neutralized(
        self, sig_lite_fixture: Path, tmp_path: Path
    ) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        q = parsed.questions[0]
        ans = Answer(
            question_id=q.id,
            sheet=q.sheet,
            row=q.row,
            column=q.column,
            text="=SUM(A1:A2)",
            confidence=0.9,
        )
        out_path = tmp_path / "f.xlsx"
        assemble_filled_xlsx([ans], str(sig_lite_fixture), str(out_path))
        wb = load_workbook(str(out_path))
        ws = wb.active
        cell = ws.cell(row=ans.row, column=ans.column)
        assert isinstance(cell.value, str)
        assert not cell.value.startswith("=")

    def test_round_trip_parses_back(
        self, sig_lite_fixture: Path, tmp_path: Path
    ) -> None:
        parsed = parse_xlsx(str(sig_lite_fixture))
        answers = [
            Answer(
                question_id=q.id,
                sheet=q.sheet,
                row=q.row,
                column=q.column,
                text="Yes",
                confidence=0.9,
            )
            for q in parsed.questions
        ]
        out_path = tmp_path / "rt.xlsx"
        assemble_filled_xlsx(answers, str(sig_lite_fixture), str(out_path))
        # Re-parse the filled XLSX — questions should still parse.
        reparsed = parse_xlsx(str(out_path))
        assert reparsed.question_count == parsed.question_count

    def test_missing_source_raises(self, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            assemble_filled_xlsx([], str(tmp_path / "missing.xlsx"), str(tmp_path / "o.xlsx"))


class TestRoundTripIntegration:
    def test_parse_cluster_metadata_assemble_pipeline(
        self, sig_lite_fixture: Path, tmp_path: Path
    ) -> None:
        """End-to-end pipeline through all four tools."""
        parsed = parse_xlsx(str(sig_lite_fixture))
        clusters = cluster_questions(parsed)
        metadata = extract_question_metadata(parsed)

        assert isinstance(parsed, ParsedQuestionnaire)
        assert clusters.cluster_count > 0
        assert metadata.metadata_count == parsed.question_count

        answers = [
            Answer(
                question_id=q.id,
                sheet=q.sheet,
                row=q.row,
                column=q.column,
                text=f"Answer to {q.id}",
                confidence=0.85,
            )
            for q in parsed.questions
        ]
        out = tmp_path / "pipeline.xlsx"
        result = assemble_filled_xlsx(answers, str(sig_lite_fixture), str(out))
        assert result.answers_written == parsed.question_count
        assert result.flagged_count == 0
