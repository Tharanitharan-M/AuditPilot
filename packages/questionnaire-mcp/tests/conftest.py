"""Pytest fixtures for questionnaire-mcp."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook


def _section_row(sheet, row: int, label: str) -> None:
    """Write a single-cell section heading row."""
    sheet.cell(row=row, column=1, value=label)


def _question_row(sheet, row: int, qid: str, text: str) -> None:
    """Write a SIG-Lite-style question row.

    Columns: 1 = id, 2 = question, 3 = answer.
    """
    sheet.cell(row=row, column=1, value=qid)
    sheet.cell(row=row, column=2, value=text)


@pytest.fixture
def sig_lite_fixture(tmp_path: Path) -> Path:
    """A small but representative SIG-Lite-shaped fixture XLSX (24 questions, 6 domains)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SIG-Lite"

    rows: list[tuple[str, list[tuple[str, str]]]] = [
        (
            "Access Control",
            [
                (
                    "A.1.1",
                    "Do you require multi-factor authentication for administrative access?",
                ),
                ("A.1.2", "Is access to production systems granted on a least privilege basis?"),
                ("A.1.3", "How are passwords stored and rotated for privileged accounts?"),
                ("A.1.4", "Does your SSO solution federate to all production systems?"),
            ],
        ),
        (
            "Data Handling",
            [
                ("D.1.1", "Is data encrypted at rest using strong cryptographic algorithms?"),
                ("D.1.2", "Do you classify data and apply DLP controls to PII?"),
                ("D.1.3", "What is your data retention policy for customer personal data?"),
                ("D.1.4", "Is encryption in transit enforced via TLS 1.2 or higher?"),
            ],
        ),
        (
            "Incident Response",
            [
                ("IR.1.1", "Do you maintain a documented incident response playbook?"),
                ("IR.1.2", "Is there a 24x7 on-call rotation for incident triage?"),
                ("IR.1.3", "How many security incidents have you had in the last 12 months?"),
                ("IR.1.4", "When did you last conduct an incident response tabletop?"),
            ],
        ),
        (
            "Network Security",
            [
                ("N.1.1", "Are firewalls configured to deny by default at all network boundaries?"),
                ("N.1.2", "Do you operate an IDS or IPS at the network perimeter?"),
                ("N.1.3", "Is your VPN configured with MFA and split tunneling disabled?"),
                ("N.1.4", "Are network segments isolated by VLAN or equivalent?"),
            ],
        ),
        (
            "Third-Party Management",
            [
                ("V.1.1", "Do you maintain a vendor inventory with risk tiers?"),
                ("V.1.2", "Are subprocessors reviewed annually for compliance?"),
                ("V.1.3", "What is your contract clause for vendor data breach notification?"),
                ("V.1.4", "Do you require third-party SOC 2 readiness reports for tier-1 vendors?"),
            ],
        ),
        (
            "Training and Awareness",
            [
                ("T.1.1", "Is security awareness training mandatory for all employees?"),
                ("T.1.2", "Do you run phishing simulation campaigns at least quarterly?"),
                ("T.1.3", "How many phishing simulations were run in the last 12 months?"),
                ("T.1.4", "Is role-based security training required for engineering staff?"),
            ],
        ),
    ]

    cur = 1
    for section, qs in rows:
        _section_row(ws, cur, section)
        cur += 1
        for qid, text in qs:
            _question_row(ws, cur, qid, text)
            cur += 1

    out = tmp_path / "sig_lite_fixture.xlsx"
    wb.save(str(out))
    return out
