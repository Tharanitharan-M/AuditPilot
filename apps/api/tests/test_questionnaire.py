"""Sprint 7 chunks 7.6, 7.7, 7.8, 7.10, 7.11 — questionnaire workspace tests.

Covers:
  - POST /api/questionnaire/upload validates size, type, persists run, enqueues job.
  - GET /api/questionnaire lists user's runs.
  - GET /api/questionnaire/{run_id} returns run + questions.
  - PATCH /api/questionnaire/questions/{id} clears flag and persists edit.
  - GET /api/questionnaire/{run_id}/poll returns the live status row.
  - GET /api/questionnaire/{run_id}/download 302s when ready, 409s otherwise.
  - QuestionnaireFillHandler: end-to-end pipeline against an in-memory pool.
"""

from __future__ import annotations

import json
import uuid
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from apps.api.auth.clerk import ClerkUser

FAKE_USER = ClerkUser(user_id="user_q1", session_id="sess_q1")
OTHER_USER = ClerkUser(user_id="user_q2", session_id="sess_q2")


# ── App harness ──────────────────────────────────────────────────────────────


def _make_client(
    monkeypatch: pytest.MonkeyPatch, user: ClerkUser = FAKE_USER
) -> TestClient:
    monkeypatch.setenv("ENVIRONMENT", "development")
    monkeypatch.setenv("DATABASE_URL", "postgres://test:test@localhost/test")
    monkeypatch.setenv("REDIS_URL", "redis://localhost:6379/0")
    monkeypatch.setenv("CLERK_SECRET_KEY", "sk_test_fake")
    monkeypatch.setenv("CLERK_PUBLISHABLE_KEY", "pk_test_fake")
    monkeypatch.setenv("GEMINI_API_KEY", "fake-gemini-key")
    monkeypatch.setenv("LANGFUSE_PUBLIC_KEY", "pk-lf-fake")
    monkeypatch.setenv("LANGFUSE_SECRET_KEY", "sk-lf-fake")
    monkeypatch.setenv("QUESTIONNAIRE_RATE_LIMIT", "1000/minute")

    import importlib

    import apps.api.main as main_module

    main_module = importlib.reload(main_module)

    from apps.api.auth.clerk import verify_clerk_token

    main_module.app.dependency_overrides[verify_clerk_token] = lambda: user
    return TestClient(main_module.app)


# ── In-memory fakes for pool, storage, job queue ─────────────────────────────


class _FakeCursor:
    def __init__(self, parent: _FakePool) -> None:
        self._parent = parent
        self._rows: list[tuple] = []

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return False

    async def execute(self, query: str, params: tuple = ()) -> None:
        await self._parent._exec(query, params)
        self._rows = list(self._parent._last_rows)

    async def fetchone(self):
        return self._rows[0] if self._rows else None

    async def fetchall(self):
        return list(self._rows)


class _FakeConn:
    def __init__(self, parent: _FakePool) -> None:
        self._parent = parent

    async def __aenter__(self):
        return self

    async def __aexit__(self, exc_type, exc, tb):
        return False

    def cursor(self):
        return _FakeCursor(self._parent)

    def transaction(self):
        return self

    async def execute(self, query: str, params: tuple = ()) -> None:
        await self._parent._exec(query, params)

    async def commit(self):
        return None

    def notifies(self):
        async def _empty():
            if False:
                yield  # pragma: no cover
        return _empty()


class _FakePool:
    """In-memory store covering questionnaire_runs and questionnaire_questions."""

    def __init__(self) -> None:
        self.runs: list[dict[str, Any]] = []
        self.questions: list[dict[str, Any]] = []
        self._last_rows: list[tuple] = []
        self.user_scope: str | None = None

    def connection(self):
        return _FakeConn(self)

    # ── seed helpers ─────────────────────────────────────────────────────────
    def seed_run(self, **fields: Any) -> str:
        run_id = fields.pop("id", str(uuid.uuid4()))
        now = datetime.now(UTC)
        record = {
            "id": run_id,
            "user_id": fields.pop("user_id", FAKE_USER.user_id),
            "filename": fields.pop("filename", "sig.xlsx"),
            "format": fields.pop("format", "sig-lite"),
            "status": fields.pop("status", "queued"),
            "source_r2_key": fields.pop("source_r2_key", "key"),
            "output_r2_key": fields.pop("output_r2_key", None),
            "question_count": fields.pop("question_count", 0),
            "answered_count": fields.pop("answered_count", 0),
            "flagged_count": fields.pop("flagged_count", 0),
            "cluster_count": fields.pop("cluster_count", 0),
            "job_idempotency_key": fields.pop("job_idempotency_key", "k"),
            "failure_reason": fields.pop("failure_reason", None),
            "created_at": fields.pop("created_at", now),
            "updated_at": fields.pop("updated_at", now),
        }
        if fields:
            raise AssertionError(f"unexpected seed fields: {fields}")
        self.runs.append(record)
        return run_id

    def seed_question(self, run_id: str, **fields: Any) -> str:
        qpk = fields.pop("id", str(uuid.uuid4()))
        record = {
            "id": qpk,
            "run_id": run_id,
            "user_id": fields.pop("user_id", FAKE_USER.user_id),
            "question_id": fields.pop("question_id", "Q1"),
            "sheet": fields.pop("sheet", "SIG"),
            "row_idx": fields.pop("row", 1),
            "column_idx": fields.pop("column", 2),
            "section": fields.pop("section", "Access Control"),
            "domain": fields.pop("domain", "access_control"),
            "answer_type": fields.pop("answer_type", "yes_no"),
            "question_text": fields.pop("question_text", "Do you require MFA?"),
            "answer_text": fields.pop("answer_text", "Yes"),
            "confidence": fields.pop("confidence", 0.9),
            "flagged": fields.pop("flagged", False),
            "citations": fields.pop("citations", []),
            "edited_by_user": fields.pop("edited_by_user", False),
        }
        if fields:
            raise AssertionError(f"unexpected seed fields: {fields}")
        self.questions.append(record)
        return qpk

    # ── SQL emulator ─────────────────────────────────────────────────────────
    async def _exec(self, query: str, params: tuple) -> None:
        q = " ".join(query.split())
        self._last_rows = []
        if "set_config" in q:
            self.user_scope = params[0]
            return
        if q.startswith("SELECT id::text FROM questionnaire_runs WHERE user_id"):
            user_id, idem = params
            hits = [
                r
                for r in self.runs
                if r["user_id"] == user_id and r["job_idempotency_key"] == idem
            ]
            hits.sort(key=lambda r: r["created_at"], reverse=True)
            self._last_rows = [(r["id"],) for r in hits[:1]]
            return
        if q.startswith("INSERT INTO questionnaire_runs"):
            run_id, user_id, filename, fmt, source_r2_key, idem = params
            now = datetime.now(UTC)
            self.runs.append(
                {
                    "id": run_id,
                    "user_id": user_id,
                    "filename": filename,
                    "format": fmt,
                    "status": "queued",
                    "source_r2_key": source_r2_key,
                    "output_r2_key": None,
                    "question_count": 0,
                    "answered_count": 0,
                    "flagged_count": 0,
                    "cluster_count": 0,
                    "job_idempotency_key": idem,
                    "failure_reason": None,
                    "created_at": now,
                    "updated_at": now,
                }
            )
            return
        if q.startswith("SELECT id::text, user_id, filename, format, status,"):
            if "AND id = %s::uuid" in q:
                user_id, run_id = params
                hits = [
                    r
                    for r in self.runs
                    if r["user_id"] == user_id and r["id"] == run_id
                ]
            else:
                user_id = params[0]
                hits = [r for r in self.runs if r["user_id"] == user_id]
                hits.sort(key=lambda r: r["created_at"], reverse=True)
            self._last_rows = [_run_row(r) for r in hits]
            return
        if q.startswith("SELECT id::text, run_id::text, question_id"):
            user_id, run_id = params
            qs = [
                r
                for r in self.questions
                if r["user_id"] == user_id and r["run_id"] == run_id
            ]
            qs.sort(key=lambda r: (r["sheet"], r["row_idx"], r["column_idx"]))
            self._last_rows = [_question_row(r) for r in qs]
            return
        if q.startswith("UPDATE questionnaire_runs SET status = 'ready', output_r2_key"):
            output_key, answered, flagged, user_id, run_id = params
            for r in self.runs:
                if r["user_id"] == user_id and r["id"] == run_id:
                    r["status"] = "ready"
                    r["output_r2_key"] = output_key
                    r["answered_count"] = answered
                    r["flagged_count"] = flagged
                    r["updated_at"] = datetime.now(UTC)
            return
        if q.startswith("UPDATE questionnaire_runs SET status = %s, updated_at"):
            new_status, user_id, run_id = params
            for r in self.runs:
                if r["user_id"] == user_id and r["id"] == run_id:
                    r["status"] = new_status
                    r["updated_at"] = datetime.now(UTC)
            return
        if q.startswith("UPDATE questionnaire_runs SET status = 'failed'"):
            reason, user_id, run_id = params
            for r in self.runs:
                if r["user_id"] == user_id and r["id"] == run_id:
                    r["status"] = "failed"
                    r["failure_reason"] = reason
                    r["updated_at"] = datetime.now(UTC)
            return
        if q.startswith("UPDATE questionnaire_runs SET question_count"):
            (
                question_count,
                cluster_count,
                answered,
                flagged,
                user_id,
                run_id,
            ) = params
            for r in self.runs:
                if r["user_id"] == user_id and r["id"] == run_id:
                    r["question_count"] = question_count
                    r["cluster_count"] = cluster_count
                    r["answered_count"] = answered
                    r["flagged_count"] = flagged
                    r["updated_at"] = datetime.now(UTC)
            return
        if q.startswith(
            "DELETE FROM questionnaire_questions WHERE user_id"
        ):
            user_id, run_id = params
            self.questions = [
                r
                for r in self.questions
                if not (r["user_id"] == user_id and r["run_id"] == run_id)
            ]
            return
        if q.startswith("INSERT INTO questionnaire_questions"):
            (
                pk,
                run_id,
                user_id,
                qid,
                sheet,
                row_idx,
                col_idx,
                section,
                domain,
                ans_type,
                qtext,
                atext,
                conf,
                flagged,
                cit_json,
            ) = params
            try:
                cit = json.loads(cit_json)
            except (TypeError, json.JSONDecodeError):
                cit = []
            self.questions.append(
                {
                    "id": pk,
                    "run_id": run_id,
                    "user_id": user_id,
                    "question_id": qid,
                    "sheet": sheet,
                    "row_idx": row_idx,
                    "column_idx": col_idx,
                    "section": section,
                    "domain": domain,
                    "answer_type": ans_type,
                    "question_text": qtext,
                    "answer_text": atext,
                    "confidence": conf,
                    "flagged": flagged,
                    "citations": cit,
                    "edited_by_user": False,
                }
            )
            return
        if q.startswith("UPDATE questionnaire_questions"):
            # Generic patch handler — emulate the dynamic field set built by route.
            params_list = list(params)
            user_id = params_list[-2]
            qpk = params_list[-1]
            for record in self.questions:
                if record["user_id"] == user_id and record["id"] == qpk:
                    record["answer_text"] = params_list[0]
                    record["edited_by_user"] = True
                    idx = 1
                    if "citations = %s::jsonb" in q:
                        try:
                            record["citations"] = json.loads(params_list[idx])
                        except (TypeError, json.JSONDecodeError):
                            pass
                        idx += 1
                    if "confidence = %s" in q:
                        record["confidence"] = params_list[idx]
                        idx += 1
                    if "flagged = FALSE" in q:
                        record["flagged"] = False
                    self._last_rows = [_question_row(record)]
                    return
            self._last_rows = []
            return
        raise AssertionError(f"unexpected SQL: {q!r}")


def _run_row(r: dict[str, Any]) -> tuple:
    return (
        r["id"],
        r["user_id"],
        r["filename"],
        r["format"],
        r["status"],
        r["question_count"],
        r["answered_count"],
        r["flagged_count"],
        r["cluster_count"],
        r["output_r2_key"],
        r["failure_reason"],
        r["created_at"],
        r["updated_at"],
    )


def _question_row(r: dict[str, Any]) -> tuple:
    return (
        r["id"],
        r["run_id"],
        r["question_id"],
        r["sheet"],
        r["row_idx"],
        r["column_idx"],
        r["section"],
        r["domain"],
        r["answer_type"],
        r["question_text"],
        r["answer_text"],
        r["confidence"],
        r["flagged"],
        r["citations"],
        r["edited_by_user"],
    )


class _FakeJobQueue:
    def __init__(self) -> None:
        self.enqueued: list[Any] = []

    async def enqueue(self, message):  # type: ignore[no-untyped-def]
        self.enqueued.append(message)
        from apps.api.jobs.schemas import JobResult

        return JobResult(message_id="mid_1", deduplicated=False)


class _FakeStorage:
    def __init__(self, backend: str = "local") -> None:
        self.put_calls: list[tuple[str, int]] = []
        self._backend = backend

    @property
    def backend(self) -> str:
        return self._backend

    def make_key(self, *, user_id: str, kind: str, suffix: str = "") -> str:
        return f"users/{user_id}/{kind}/test{suffix}"

    def put_bytes(self, key, body, *, content_type):
        self.put_calls.append((key, len(body)))
        from apps.api.services.object_storage import StoredObject

        return StoredObject(key=key, backend="local", size_bytes=len(body))

    def get_bytes(self, key):
        return b""

    def local_path(self, key):  # pragma: no cover — not used by route tests
        from pathlib import Path

        return Path(f"/tmp/{key}")

    def presigned_get_url(self, key, *, ttl_seconds: int = 900) -> str:
        return f"https://example.com/{key}?expires={ttl_seconds}"


# ── XLSX fixture ─────────────────────────────────────────────────────────────


def _build_xlsx_bytes(question_count: int = 4) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "SIG-Lite"
    ws.cell(row=1, column=1, value="Access Control")
    for i in range(question_count):
        ws.cell(row=i + 2, column=1, value=f"A.{i+1}")
        ws.cell(
            row=i + 2,
            column=2,
            value=f"Do you require multi-factor authentication question {i+1}?",
        )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Tests for /api/questionnaire/upload ──────────────────────────────────────


def test_upload_persists_run_and_enqueues(monkeypatch: pytest.MonkeyPatch) -> None:
    client = _make_client(monkeypatch)
    pool = _FakePool()
    queue = _FakeJobQueue()
    storage = _FakeStorage()

    import apps.api.main as main_module
    from apps.api.db import get_pool
    from apps.api.routes.questionnaire import _get_job_queue, _get_storage

    main_module.app.dependency_overrides[get_pool] = lambda: pool
    main_module.app.dependency_overrides[_get_job_queue] = lambda: queue
    main_module.app.dependency_overrides[_get_storage] = lambda: storage

    body = _build_xlsx_bytes(4)
    r = client.post(
        "/api/questionnaire/upload",
        headers={"Authorization": "Bearer fake"},
        files={
            "file": (
                "sig.xlsx",
                body,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 200, r.text
    out = r.json()
    assert out["status"] == "queued"
    assert out["filename"] == "sig.xlsx"
    assert out["size_bytes"] == len(body)
    assert len(pool.runs) == 1
    assert pool.runs[0]["status"] == "queued"
    assert len(queue.enqueued) == 1
    msg = queue.enqueued[0]
    assert msg.payload["run_id"] == pool.runs[0]["id"]


def test_upload_rejects_oversized_file(monkeypatch: pytest.MonkeyPatch) -> None:
    client = _make_client(monkeypatch)
    pool = _FakePool()
    queue = _FakeJobQueue()
    storage = _FakeStorage()

    import apps.api.main as main_module
    from apps.api.db import get_pool
    from apps.api.routes.questionnaire import _get_job_queue, _get_storage

    main_module.app.dependency_overrides[get_pool] = lambda: pool
    main_module.app.dependency_overrides[_get_job_queue] = lambda: queue
    main_module.app.dependency_overrides[_get_storage] = lambda: storage

    blob = b"\x50\x4b" + b"\x00" * (10 * 1024 * 1024 + 1024)  # > 10 MB
    r = client.post(
        "/api/questionnaire/upload",
        headers={"Authorization": "Bearer fake"},
        files={
            "file": (
                "big.xlsx",
                blob,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 413


def test_upload_rejects_bad_mime(monkeypatch: pytest.MonkeyPatch) -> None:
    client = _make_client(monkeypatch)
    pool = _FakePool()
    queue = _FakeJobQueue()
    storage = _FakeStorage()

    import apps.api.main as main_module
    from apps.api.db import get_pool
    from apps.api.routes.questionnaire import _get_job_queue, _get_storage

    main_module.app.dependency_overrides[get_pool] = lambda: pool
    main_module.app.dependency_overrides[_get_job_queue] = lambda: queue
    main_module.app.dependency_overrides[_get_storage] = lambda: storage

    r = client.post(
        "/api/questionnaire/upload",
        headers={"Authorization": "Bearer fake"},
        files={"file": ("notes.pdf", b"%PDF-1.4...", "application/pdf")},
    )
    assert r.status_code == 422


# ── Tests for list / get / poll ──────────────────────────────────────────────


def test_list_runs(monkeypatch: pytest.MonkeyPatch) -> None:
    client = _make_client(monkeypatch)
    pool = _FakePool()
    pool.seed_run(filename="A.xlsx")
    pool.seed_run(filename="B.xlsx", status="ready")
    pool.seed_run(user_id=OTHER_USER.user_id, filename="other.xlsx")

    import apps.api.main as main_module
    from apps.api.db import get_pool

    main_module.app.dependency_overrides[get_pool] = lambda: pool

    r = client.get(
        "/api/questionnaire", headers={"Authorization": "Bearer fake"}
    )
    assert r.status_code == 200
    body = r.json()
    assert body["count"] == 2
    names = sorted(rs["filename"] for rs in body["runs"])
    assert names == ["A.xlsx", "B.xlsx"]


def test_get_run_returns_questions(monkeypatch: pytest.MonkeyPatch) -> None:
    client = _make_client(monkeypatch)
    pool = _FakePool()
    run_id = pool.seed_run(filename="A.xlsx")
    pool.seed_question(run_id=run_id, question_text="Q1?")
    pool.seed_question(run_id=run_id, question_text="Q2?", flagged=True, confidence=0.4)

    import apps.api.main as main_module
    from apps.api.db import get_pool

    main_module.app.dependency_overrides[get_pool] = lambda: pool

    r = client.get(
        f"/api/questionnaire/{run_id}", headers={"Authorization": "Bearer fake"}
    )
    assert r.status_code == 200
    body = r.json()
    assert body["run"]["filename"] == "A.xlsx"
    assert len(body["questions"]) == 2


def test_poll_returns_run(monkeypatch: pytest.MonkeyPatch) -> None:
    client = _make_client(monkeypatch)
    pool = _FakePool()
    run_id = pool.seed_run(status="drafting", question_count=10, answered_count=4)

    import apps.api.main as main_module
    from apps.api.db import get_pool

    main_module.app.dependency_overrides[get_pool] = lambda: pool

    r = client.get(
        f"/api/questionnaire/{run_id}/poll", headers={"Authorization": "Bearer fake"}
    )
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "drafting"
    assert body["answered_count"] == 4


# ── Tests for PATCH /api/questionnaire/questions/{id} ────────────────────────


def test_patch_question_clears_flag(monkeypatch: pytest.MonkeyPatch) -> None:
    client = _make_client(monkeypatch)
    pool = _FakePool()
    run_id = pool.seed_run()
    qpk = pool.seed_question(
        run_id=run_id, flagged=True, confidence=0.4, answer_text="Pending"
    )

    import apps.api.main as main_module
    from apps.api.db import get_pool

    main_module.app.dependency_overrides[get_pool] = lambda: pool

    r = client.patch(
        f"/api/questionnaire/questions/{qpk}",
        headers={"Authorization": "Bearer fake"},
        json={
            "answer_text": "Yes — see SOC 2 readiness reference architecture.",
            "clear_flag": True,
        },
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["flagged"] is False
    assert body["edited_by_user"] is True
    assert "readiness" in body["answer_text"]


# ── Tests for download ───────────────────────────────────────────────────────


def test_download_redirects_when_ready(monkeypatch: pytest.MonkeyPatch) -> None:
    client = _make_client(monkeypatch)
    pool = _FakePool()
    storage = _FakeStorage(backend="r2")
    run_id = pool.seed_run(status="ready", output_r2_key="users/x/q/out.xlsx")

    import apps.api.main as main_module
    from apps.api.db import get_pool
    from apps.api.routes.questionnaire import _get_storage

    main_module.app.dependency_overrides[get_pool] = lambda: pool
    main_module.app.dependency_overrides[_get_storage] = lambda: storage

    r = client.get(
        f"/api/questionnaire/{run_id}/download",
        headers={"Authorization": "Bearer fake"},
        follow_redirects=False,
    )
    assert r.status_code == 302
    assert "out.xlsx" in r.headers["location"]


def test_download_409_when_not_ready(monkeypatch: pytest.MonkeyPatch) -> None:
    client = _make_client(monkeypatch)
    pool = _FakePool()
    storage = _FakeStorage()
    run_id = pool.seed_run(status="drafting")

    import apps.api.main as main_module
    from apps.api.db import get_pool
    from apps.api.routes.questionnaire import _get_storage

    main_module.app.dependency_overrides[get_pool] = lambda: pool
    main_module.app.dependency_overrides[_get_storage] = lambda: storage

    r = client.get(
        f"/api/questionnaire/{run_id}/download",
        headers={"Authorization": "Bearer fake"},
        follow_redirects=False,
    )
    assert r.status_code == 409


# ── Worker handler integration ───────────────────────────────────────────────


@pytest.mark.asyncio
async def test_handler_pipeline_marks_ready(
    monkeypatch: pytest.MonkeyPatch, tmp_path
) -> None:
    """End-to-end: handler parses an XLSX, writes questions, marks run ready."""
    from pathlib import Path

    from apps.api.jobs.schemas import JobMessage, JobType
    from apps.api.services.object_storage import ObjectStorage, reset_object_storage
    from apps.api.services.questionnaire_worker import QuestionnaireFillHandler

    monkeypatch.setenv("LOCAL_OBJECT_STORAGE_DIR", str(tmp_path))
    reset_object_storage()

    pool = _FakePool()

    # Place a real XLSX in the local storage backend so the handler can
    # parse + assemble it.
    xlsx_bytes = _build_xlsx_bytes(4)
    src_key = f"users/{FAKE_USER.user_id}/questionnaires/seed.xlsx"
    src_path = tmp_path / src_key
    src_path.parent.mkdir(parents=True, exist_ok=True)
    src_path.write_bytes(xlsx_bytes)

    # Set up the run row before kicking off the handler (simulating /upload).
    run_id = pool.seed_run(source_r2_key=src_key, status="queued")

    from apps.api.config import Settings

    settings = Settings(
        environment="development",
        clerk_secret_key="sk_test_fake",
        clerk_publishable_key="pk_test_fake",
    )
    storage = ObjectStorage(settings)
    handler = QuestionnaireFillHandler(
        pool_factory=lambda: pool,  # type: ignore[arg-type]
        storage=storage,
    )

    msg = JobMessage(
        type=JobType.QUESTIONNAIRE_FILL,
        user_id=FAKE_USER.user_id,
        idempotency_key="qf:test:1",
        payload={"run_id": run_id, "source_r2_key": src_key},
    )
    await handler(msg)

    finished = next(r for r in pool.runs if r["id"] == run_id)
    assert finished["status"] == "ready"
    assert finished["question_count"] == 4
    assert finished["answered_count"] == 4
    assert finished["output_r2_key"] is not None
    out_path = Path(tmp_path) / finished["output_r2_key"]
    assert out_path.exists()
    qs = [q for q in pool.questions if q["run_id"] == run_id]
    assert len(qs) == 4
